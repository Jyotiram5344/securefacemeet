"""CSV / XLSX / PDF builders for participant attendance exports."""
from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

EXPORT_KEYS = [
    "meeting_id",
    "log_id",
    "user_id",
    "full_name",
    "email",
    "student_external_id",
    "student_class",
    "join_time",
    "exit_time",
    "seconds_present",
    "scheduled_duration_seconds",
    "dwell_ratio",
    "dwell_percent",
    "meets_dwell_threshold",
    "meets_face_threshold",
    "status",
    "fully_qualified",
]


def to_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    si = StringIO()
    si.write("\ufeff")
    w = csv.writer(si)
    if not rows:
        w.writerow(["(no participant rows in this export)"])
        return si.getvalue().encode("utf-8")
    w.writerow(EXPORT_KEYS)
    for raw in rows:
        w.writerow([raw.get(k, "") for k in EXPORT_KEYS])
    return si.getvalue().encode("utf-8")


def to_xlsx_bytes(rows: list[dict[str, Any]], sheet_title: str = "Participants") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Participants"
    if not rows:
        ws.append(["(no participant rows in this export)"])
    else:
        ws.append(EXPORT_KEYS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for raw in rows:
            ws.append([raw.get(k, "") for k in EXPORT_KEYS])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class _ParticipantPDF(FPDF):
    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def to_pdf_bytes(rows: list[dict[str, Any]], title: str) -> bytes:
    pdf = _ParticipantPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.multi_cell(0, 7, title)
    pdf.ln(2)

    if not rows:
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, "No participant rows in this export.", ln=True)
        return bytes(pdf.output())

    keys = EXPORT_KEYS
    col_w = max(35, min(267 / len(keys), 48))
    pdf.set_font("Helvetica", "B", 7)
    for k in keys:
        pdf.cell(col_w, 6, str(k)[:26], border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for raw in rows:
        for k in keys:
            v = raw.get(k, "")
            if isinstance(v, bool):
                v = "yes" if v else "no"
            txt = str(v)[:32]
            pdf.cell(col_w, 5.5, txt, border=1)
        pdf.ln()

    return bytes(pdf.output())
